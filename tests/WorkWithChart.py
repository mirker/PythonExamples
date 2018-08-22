#!python

import openpyxl
from openpyxl.chart import BarChart, Reference, Series


wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, 11):
	sheet['A'+str(i)] = i
	
refObj = Reference(sheet, min_col=1, min_row=1, max_row=10, max_col=1)
#seriesObj = Series(refObj, title='First series')
cats = Reference(sheet, 1, 1, 10)
chartObj = BarChart()
#chartObj.type = "col"
#chartObj.style = 10
#chartObj.title = "Bar Chart"
#chartObj.y_axis.title = 'Test number'
#chartObj.x_axis.title = 'Sample length (mm)'
chartObj.add_data(refObj)
#chartObj.set_categories(cats)
#chartObj.drawing.top = 50
#chartObj.drawing.left = 100
#chartObj.drawing.width = 300
#chartObj.drawing.height = 200
#chartObj.shape = 4
sheet.add_chart(chartObj)
wb.save('myChart.xlsx')
'''
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()

rows = [
    ('Number', 'Batch 1', 'Batch 2'),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]


for row in rows:
    ws.append(row)


chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A10")

wb.save('myChart.xlsx')
'''